import csv
import re
import openpyxl
import os
import shutil


def concat_values_ab(file_path):
    try:
        csv.field_size_limit(10**9)
        with open(file_path, 'r', newline='') as csvfile:
            csv_reader = csv.reader(csvfile)
            rows = list(csv_reader)

        pattern = re.compile(r'\((\d+\s*)+\)')

        for row_index, row in enumerate(rows):
            for col_index, cell in enumerate(row):
                rows[row_index][col_index] = pattern.sub(lambda x: x.group().replace(' ', ''), cell)

        with open(file_path, 'w', newline='') as csvfile:
            csv_writer = csv.writer(csvfile)
            csv_writer.writerows(rows)

    except FileNotFoundError:
        print("File not found.")

# Example usage:
file_path = "./output1.csv"
concat_values_ab(file_path)

def extract_patterns_from_csv(file_path, pattern):
    try:
        csv.field_size_limit(10**9)
        all_matches = []
        with open(file_path, "r", newline='', encoding='utf-8') as csv_file:
            csv.field_size_limit(1000000)
            csv_reader = csv.reader(csv_file)
            for row in csv_reader:
                for cell in row:
                    matches = re.finditer(pattern, cell)
                    for match in matches:
                        matched_value = match.group()  # Get the matched value
                        all_matches.append((matched_value, match.group(0)))  # Append value and pattern
        return all_matches
    except csv.Error as e:
        print('CSV file error:', e)
        return []

def replace_values_in_csv(input_file, output_file, pattern):
    try:
        csv.field_size_limit(10**9)
        with open(input_file, "r", newline='', encoding='utf-8') as csv_file:
            csv_reader = csv.reader(csv_file)
            rows = list(csv_reader)

        with open(output_file, "w", newline='', encoding='utf-8') as csv_file:
            csv_writer = csv.writer(csv_file)
            for row in rows:
                updated_row = [re.sub(pattern, lambda x: x.group().replace(' ', ''), cell) for cell in row]
                csv_writer.writerow(updated_row)

        # print("Replacement completed. Updated content:")
        with open(output_file, "r", newline='', encoding='utf-8') as csv_file:
            csv_file.read()
    except csv.Error as e:
        print('CSV file error:', e)

# Define the CSV file paths
csv_input_file = "output1.csv"
csv_output_file = "output1.csv"

# Define the regex pattern
pattern = r'\b(\d{1})\s(\d{2})\b'
reverse_pattern = r'\b(\d{2})\s(\d{1})\b'

# Perform replacements in the CSV file
replace_values_in_csv(csv_input_file, csv_output_file, pattern)
replace_values_in_csv('output.csv', 'output.csv', pattern)
# replace_values_in_csv(csv_input_file, csv_output_file, reverse_pattern)
# replace_values_in_csv('output.csv', 'output.csv', reverse_pattern)

def fix_values_in_csv(input_file):
    try:
        csv.field_size_limit(10**9)
        with open(input_file, "r", newline='', encoding='utf-8') as csv_file:
            csv_reader = csv.reader(csv_file)
            rows = list(csv_reader)

        # Create a new list to store the updated rows
        updated_rows = []
        for row in rows:
            updated_row = []
            for cell in row:
                # Remove spaces and concatenate the values based on the specified patterns
                updated_cell = re.sub(r'\b(\d)\s(\d[A-Z])\b', r'\1\2', cell)
                updated_cell = re.sub(r'\b(\d{2})\s([A-Z])\b', r'\1\2', updated_cell)

                updated_row.append(updated_cell)
            # Append the updated row to the list of updated rows
            updated_rows.append(updated_row)

        # Write the updated rows back to the CSV file
        with open(input_file, "w", newline='', encoding='utf-8') as csv_file:
            csv_writer = csv.writer(csv_file)
            csv_writer.writerows(updated_rows)

        print("Values fixed in the CSV file.")
    except csv.Error as e:
        print('CSV file error:', e)

# Define the CSV file path
csv_input_file = "output.csv"

# Fix values in the CSV file
fix_values_in_csv(csv_input_file)
fix_values_in_csv('output1.csv')


def remove_patterns_from_csv(input_file):
    try:
        csv.field_size_limit(10**9)
        with open(input_file, "r", newline='', encoding='utf-8') as csv_file:
            csv_reader = csv.reader(csv_file)
            rows = list(csv_reader)

        with open(input_file, "w", newline='', encoding='utf-8') as csv_file:
            csv_writer = csv.writer(csv_file)
            for row in rows:
                updated_row = [re.sub(r'\b(?:RPV|PPR|P\sPR|PP\sR|RP\sV|R\sPV|ABS|A\sBS|AB\sS|RLE|R\sLE|RL\sE)\b', '', cell) for cell in row]
                csv_writer.writerow(updated_row)

        print("Patterns removed from the CSV file.")
    except csv.Error as e:
        print('CSV file error:', e)

# Define the CSV file path
csv_input_file = "output1.csv"

# Remove patterns from the CSV file
remove_patterns_from_csv(csv_input_file)
remove_patterns_from_csv('output.csv')


def concat_and_update_csv(file_path):
    try:
        # Read the CSV file
        with open(file_path, 'r', newline='') as csvfile:
            csv_reader = csv.reader(csvfile)
            rows = list(csv_reader)

        # Define the regular expression pattern
        pattern = re.compile(r'\b([A-Z])\s+([A-Z]{2})\b')

        # Process each cell in the CSV data
        for row_index, row in enumerate(rows):
            for col_index, cell in enumerate(row):
                # Replace the matched words with "PPR"
                replaced_value = pattern.sub(r'\1\2', cell)
                # Update the cell value if there was a replacement
                if replaced_value != cell:
                    rows[row_index][col_index] = replaced_value

        # Write the updated content back to the CSV file
        with open(file_path, 'w', newline='') as csvfile:
            csv_writer = csv.writer(csvfile)
            csv_writer.writerows(rows)

        print("CSV file updated successfully.")

    except FileNotFoundError:
        print("File not found.")

# Example usage:
file_path = "./output1.csv"
concat_and_update_csv(file_path)


def extract_names_from_csv(file_path):
    names_dict = {}
    csv.field_size_limit(1000000)
    with open(file_path, 'r', newline='', encoding='utf-8') as file:
        reader = csv.reader(file)
        for row in reader:
            for item in row:   
                matches = re.findall(r'(\b\d{7}\b|\b\d{8}\b)\s*(?:/)?\s*([A-Z][A-Z\s]+?)\s+\s', item)
                for match in matches:
                    seat_number, name = match
                    names_dict[seat_number] = name
    return names_dict

def remove_name_content1(input_file, output_file):
    with open(input_file, 'r', newline='', encoding='utf-8') as infile:
        file_content = infile.read()

    sections = file_content.split('\n')
    modified_sections = []
    for section in sections:
        has_seven_or_eight_digit = re.search(r'\b\d{8}\b|\b\d{7}\b', section)
        if has_seven_or_eight_digit:
            # Define the pattern to match and remove name content along with data after the name in the same line
            pattern = r'\b[A-Z][A-Z\s/]+.*'  # Pattern to match name content and data after it
            # Remove name content and data after the name from the section
            modified_section = re.sub(pattern, '', section)
        else:
            modified_section = section
        modified_sections.append(modified_section)

    modified_content = '\n'.join(modified_sections)
    
    with open(output_file, 'w', newline='', encoding='utf-8') as outfile:
        outfile.write(modified_content)

# Example usage:
remove_name_content1('./output1.csv', './output2.csv')
print("Name content and data after name removed from the CSV file.")

def replace_pattern_in_csv(file_path):
    csv.field_size_limit(10**9)
    try:
        # Read the CSV file and store its content
        with open(file_path, 'r', newline='') as csvfile:
            csv_reader = csv.reader(csvfile)
            rows = list(csv_reader)

        pattern = re.compile(r'\bP\s+P\s*([PR]?)\b|\bPP\s+R\b')

        # Process each cell and replace the matched patterns
        for row_index, row in enumerate(rows):
            for col_index, cell in enumerate(row):
                # Replace the patterns with "PPR"
                replaced_value = pattern.sub(r'PPR', cell)
                # Update the cell value if there was a replacement
                if replaced_value != cell:
                    rows[row_index][col_index] = replaced_value

        # Write the updated content back to the CSV file
        with open(file_path, 'w', newline='') as csvfile:
            csv_writer = csv.writer(csvfile)
            csv_writer.writerows(rows)

        print("Patterns replaced successfully in the CSV file.")

    except FileNotFoundError:
        print("File not found.")

# Example usage:
file_path = "output2.csv"  # Replace with the path to your CSV file
replace_pattern_in_csv(file_path)

def concatenate_patterns_in_file(file_path):
    csv.field_size_limit(1000000)
    # Create a temporary file to store the modified content
    temp_file_path = file_path + '.tmp'

    # Open the original file for reading and the temporary file for writing
    with open(file_path, 'r', newline='', encoding='utf-8') as input_file:
        with open(temp_file_path, 'w', newline='', encoding='utf-8') as temp_file:
            reader = csv.reader(input_file)
            writer = csv.writer(temp_file)

            # Process each row in the input file
            for row in reader:
                modified_row = [concatenate_patterns(pattern_string) for pattern_string in row]
                writer.writerow(modified_row)

    # Replace the original file with the modified content
    shutil.move(temp_file_path, file_path)

def concatenate_patterns(pattern_string):
    csv.field_size_limit(1000000)
    # Define the pattern to match a number followed by a space and a letter
    pattern = r'(\d{2}|\d)\s([A-Z])'
    
    # Replace the matched pattern with the concatenated value
    concatenated_string = re.sub(pattern, r'\1\2', pattern_string)
    
    return concatenated_string

# Example usage:
file_path = 'output2.csv'
concatenate_patterns_in_file(file_path)


def extract_seat_numbers_from_csv(file_path):
    seat_numbers_dict = {}
    csv.field_size_limit(1000000)
    with open(file_path, 'r', newline='', encoding='utf-8') as file:
        reader = csv.reader(file)
        for row in reader:
            # Extract seat numbers from the row
            for item in row:
                matches = re.findall(r'\b\d{8}\b|\b\d{7}\b', item)
                for match in matches:
                    seat_number = match
                    seat_numbers_dict[seat_number] = None  # Assign a placeholder value
    return seat_numbers_dict



    




def extract_marks_from_csv(file_path):
    marks_dict = {}
    updated_marks_dict = []
    csv.field_size_limit(1000000)
    with open(file_path, 'r', newline='', encoding='utf-8') as file:
        reader = csv.reader(file)
        for row in reader:
            # Join the elements of the row into a single string
            text = ' '.join(row)
            
            # # Define the marks pattern
            marks_pattern = r'--|-|\d{2}[A-Z]|\d{1}[A-Z]|\d{1}\.\d{2}|\d{2}\.\d{2}|\d\.\s\d{2}|\d\s\.\d{2}|\b\d+\.*\d*\b|\b[A-Z]\b'
            
            # Extract marks using the pattern
            marks = re.findall(marks_pattern, text)

            # pattern = r'\b\d{16}\b'
            pattern = r'\b\d{8}\b|\b\d{7}\b'

            # Find all matches of the pattern in the text
            matches = re.findall(pattern, ' '.join(marks))

            # Extract marks for each student
            for match in matches:
                index = marks.index(match)
                values = re.findall(r'--|-|\d{2}[A-Z]|\d{1}\.\d{2}|\d{2}\.\d{2}|\d\.\s\d{2}|\d\s\.\d{2}|\d{1}[A-Z]|\b\d+\b|\b[A-Z]\b', ' '.join(marks[index:]))
                marks_dict[match] = values[1:24]
                
            for key, value in marks_dict.items():
                 updated_marks_dict.append(value)

    return marks_dict , updated_marks_dict




file_path1 = './output.csv'  # changed done from output1.csv to output.csv
marks_data , updated_mark = extract_marks_from_csv(file_path1)


# print("marks data: ")
# print(updated_mark)
# from data import remove_name_content1
remove_name_content1('./output.csv', './output.csv')
def replace_values_in_csv(input_file):
    csv.field_size_limit(1000000)
    output_file = input_file  # Overwrite the input file

    # Define the patterns for finding values
    patterns = {str(i): r'\b{}\s{}\b'.format(str(i // 10), str(i % 10)) for i in range(10, 100)}


    # Read the content of the input CSV file and store it in a list
    with open(input_file, 'r', newline='') as infile:
        rows = list(csv.reader(infile))

    # Replace values based on patterns
    for pattern_key, pattern_value in patterns.items():
        for row in rows:
            for i, value in enumerate(row):
                row[i] = re.sub(pattern_value, pattern_key, value)

    # Write the modified content back to the input CSV file
    with open(input_file, 'w', newline='') as outfile:
        writer = csv.writer(outfile)
        writer.writerows(rows)

# Specify the input CSV file path
input_file = 'output.csv'

# Call the function to replace values
replace_values_in_csv(input_file)
# Define the CSV file path
csv_file_path = "output.csv"

# Define the regex pattern
pattern = r"\s{2}([0-9])\s{1}([0-9])\s{2}"

# Function to update CSV file with concatenated values
def update_csv_with_concatenated_values(csv_file_path, pattern):
    csv.field_size_limit(1000000)
    # Create a temporary file to store updated content
    temp_file_path = "temp_output.csv"
    
    # Open the CSV file and the temporary file
    with open(csv_file_path, "r", newline='', encoding='utf-8') as csv_file, \
         open(temp_file_path, "w", newline='', encoding='utf-8') as temp_file:
        csv_reader = csv.reader(csv_file)
        csv_writer = csv.writer(temp_file)

        # Iterate over each row in the CSV file
        for row in csv_reader:
            updated_row = []
            # Iterate over each cell in the row
            for cell in row:
                # Replace matched substrings with concatenated values
                updated_cell = re.sub(pattern, lambda match: match.group(1) + match.group(2), cell)
                updated_row.append(updated_cell)
            # Write the updated row to the temporary file
            csv_writer.writerow(updated_row)

    # Replace the original CSV file with the temporary file
    os.replace(temp_file_path, csv_file_path)

# Update the CSV file with concatenated values
update_csv_with_concatenated_values(csv_file_path, pattern)

marks_dict = {}
updated_marks_dict = []
def extract_marks1_from_csv(file_path):
    csv.field_size_limit(1000000)
    with open(file_path, 'r', newline='', encoding='utf-8') as file:
        reader = csv.reader(file)
        for row in reader:
            # Join the elements of the row into a single string
            text = ' '.join(row)
            
            # # Define the marks pattern
            marks_pattern = r'--|-|\d{2}[A-Z]|\d{1}[A-Z]|\d{1}\.\d{2}|\d{2}\.\d{2}|\d\.\s\d{2}|\d\s\.\d{2}|\b\d+\.*\d*\b|\d{1}$|\b[A-Z]\b'
            
            # Extract marks using the pattern
            marks = re.findall(marks_pattern, text)

            # pattern = r'\b\d{16}\b'
            pattern = r'\b\d{8}\b|\b\d{7}\b'

            # Find all matches of the pattern in the text
            matches = re.findall(pattern, ' '.join(marks))

            # Extract marks for each student
            for match in matches:
                index = marks.index(match)
                values = re.findall(r'--|-|\d{2}[A-Z]|\d{1}\.\d{2}|\d{2}\.\d{2}|\d\.\s\d{2}|\d\s\.\d{2}|\d{1}[A-Z]|\b\d+\b|\d{1}$|\b[A-Z]\b', ' '.join(marks[index:]))
                marks_dict[match] = values[1:55]
                
            for key, value in marks_dict.items():
                 updated_marks_dict.append(value)

    return marks_dict , updated_marks_dict



def extract_point_values_from_csv(elements_flat):
    file_path = './output.csv'
    text, elements_flat = extract_marks1_from_csv(file_path)
    all_values = []
    point_index = None
    # Define different patterns for the point value
    patterns = [
        r'\d\s\.\d{2}',  # Pattern: digit, space, period, two digits
        r'\d\.\s\d{2}',
        r'\d{1}\.\d{2}',  # Pattern: one digit, period, two digits
        r'--'  # Pattern: -- (hyphen hyphen)
    ]

    # Iterate over each sublist in elements_flat
    for sublist in elements_flat:
        # Iterate through the elements in reverse order
        for i in range(len(sublist) - 1, -1, -1):
            element = sublist[i]
            # Iterate through patterns to find a match
            for pattern in patterns:
                match = re.search(pattern, element)
                if match:
                    point_index = i
                    break
            if match:
                break

        # If a point value is found, store the values before the point value
        if point_index is not None:
            values_before_point = sublist[max(point_index - 30, 0):point_index + 1]
            all_values.append(values_before_point)
        else:
            all_values.append([])  # Append an empty list if no point value is found
            
    return all_values



# Create a new Excel workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

def write_marks_to_excel(ws, marks_data):
    for row_index, marks_row in enumerate(marks_data, start=1):
        for col_index, mark in enumerate(marks_row, start=3):
            ws.cell(row=row_index, column=col_index, value=mark)

def write_marks1_to_excel(ws, point_values):
    for row_index, marks_row in enumerate(point_values, start=1):
        for col_index, point_values in enumerate(marks_row, start=26):
            ws.cell(row=row_index, column=col_index, value=point_values)

def write_seat_numbers_to_excel(ws, seat_numbers):
    for row_index, seat_number in enumerate(seat_numbers, start=1):
        ws.cell(row=row_index, column=1, value=seat_number)

def write_names_to_excel(ws, names_data):
    for row_index, (seat_number, name) in enumerate(names_data.items(), start=1):
        ws.cell(row=row_index, column=2, value=name)



def add_data_to_excel(point_values,marks_data, seat_numbers, names_data, output_excel):
    wb = openpyxl.Workbook()
    ws = wb.active

    write_marks1_to_excel(ws,point_values)
    write_marks_to_excel(ws, marks_data)
    write_seat_numbers_to_excel(ws, seat_numbers)
    write_names_to_excel(ws, names_data)

    wb.save(output_excel)
    print("Data added to Excel file successfully.")

# Example usage:
file_path1 = './output.csv'
file_path2 = './output1.csv'
file_extra = './output_extra.csv'
file_seatno = 'seatno.csv'
file_path = './output.csv'

point_values = extract_point_values_from_csv(file_path)
marks_data, updated_marks_data = extract_marks_from_csv(file_path1)
names_data = extract_names_from_csv(file_extra)
seat_numbers = extract_seat_numbers_from_csv(file_seatno)

output_excel = 'output_format1.xlsx'
print("format1 excel file generated successfully")
add_data_to_excel(point_values,updated_marks_data, seat_numbers, names_data, output_excel)
